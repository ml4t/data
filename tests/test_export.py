"""Tests for export functionality."""

import json
import tempfile
from copy import deepcopy
from datetime import UTC, datetime
from pathlib import Path

import polars as pl
import pytest

from ml4t.data.core.models import DataObject, Metadata
from ml4t.data.export.formats import CSVExporter, ExcelExporter, ExportConfig, JSONExporter
from ml4t.data.export.manager import ExportManager
from ml4t.data.storage import FlatStorage, HiveStorage, StorageConfig


class TestCSVExporter:
    """Test CSV export functionality."""

    @pytest.fixture
    def sample_data(self) -> pl.DataFrame:
        """Create sample OHLCV data."""
        return pl.DataFrame(
            {
                "timestamp": ["2024-01-01", "2024-01-02", "2024-01-03"],
                "open": [100.0, 101.0, 102.0],
                "high": [105.0, 106.0, 107.0],
                "low": [99.0, 100.0, 101.0],
                "close": [104.0, 105.0, 106.0],
                "volume": [1000000, 1100000, 1200000],
            }
        )

    def test_export_csv(self, sample_data: pl.DataFrame) -> None:
        """Test basic CSV export."""
        with tempfile.TemporaryDirectory() as tmpdir:
            config = ExportConfig(
                output_path=Path(tmpdir),
                format="csv",
            )

            exporter = CSVExporter(config)
            result = exporter.export(sample_data, "TEST")

            assert result.success
            assert result.rows_exported == 3
            assert result.output_path.exists()
            assert result.output_path.name == "TEST.csv"

            # Read back and verify
            df = pl.read_csv(result.output_path)
            assert len(df) == 3
            assert df["close"][0] == 104.0

    def test_export_csv_compressed(self, sample_data: pl.DataFrame) -> None:
        """Test compressed CSV export."""
        with tempfile.TemporaryDirectory() as tmpdir:
            config = ExportConfig(
                output_path=Path(tmpdir),
                format="csv",
                compression="gzip",
            )

            exporter = CSVExporter(config)
            result = exporter.export(sample_data, "TEST")

            assert result.success
            assert result.output_path.name == "TEST.csv.gz"
            assert result.output_path.exists()

    def test_export_csv_with_transformations(self, sample_data: pl.DataFrame) -> None:
        """Test CSV export with transformations."""
        with tempfile.TemporaryDirectory() as tmpdir:
            config = ExportConfig(
                output_path=Path(tmpdir),
                format="csv",
                columns=["timestamp", "close"],
                add_returns=True,
            )

            exporter = CSVExporter(config)
            result = exporter.export(sample_data, "TEST")

            assert result.success

            # Read back and verify
            df = pl.read_csv(result.output_path)
            assert "close" in df.columns
            assert "returns" in df.columns
            assert len(df.columns) == 3  # timestamp, close, returns

    def test_date_only_filter_includes_intraday_end_date(self, tmp_path) -> None:
        data = pl.DataFrame(
            {
                "timestamp": [
                    datetime(2024, 1, 31, 16, 0, tzinfo=UTC),
                    datetime(2024, 2, 1, 9, 30, tzinfo=UTC),
                ],
                "close": [100.0, 101.0],
            }
        )
        exporter = CSVExporter(
            ExportConfig(
                output_path=tmp_path / "filtered.csv",
                format="csv",
                date_filter=("2024-01-31", "2024-01-31"),
            )
        )

        result = exporter.export(data, "AAPL")

        assert result.success, result.error
        assert pl.read_csv(result.output_path).height == 1

    def test_export_batch_csv(self, sample_data: pl.DataFrame) -> None:
        """Test batch CSV export."""
        with tempfile.TemporaryDirectory() as tmpdir:
            config = ExportConfig(
                output_path=Path(tmpdir),
                format="csv",
            )

            datasets = {
                "AAPL": sample_data,
                "GOOGL": sample_data.clone(),
            }

            exporter = CSVExporter(config)
            results = exporter.export_batch(datasets)

            assert len(results) == 2
            assert all(r.success for r in results)
            assert (Path(tmpdir) / "AAPL.csv").exists()
            assert (Path(tmpdir) / "GOOGL.csv").exists()


class TestJSONExporter:
    """Test JSON export functionality."""

    @pytest.fixture
    def sample_data(self) -> pl.DataFrame:
        """Create sample OHLCV data."""
        return pl.DataFrame(
            {
                "timestamp": ["2024-01-01", "2024-01-02", "2024-01-03"],
                "open": [100.0, 101.0, 102.0],
                "high": [105.0, 106.0, 107.0],
                "low": [99.0, 100.0, 101.0],
                "close": [104.0, 105.0, 106.0],
                "volume": [1000000, 1100000, 1200000],
            }
        )

    def test_export_json(self, sample_data: pl.DataFrame) -> None:
        """Test basic JSON export."""
        with tempfile.TemporaryDirectory() as tmpdir:
            config = ExportConfig(
                output_path=Path(tmpdir),
                format="json",
            )

            exporter = JSONExporter(config)
            result = exporter.export(sample_data, "TEST")

            assert result.success
            assert result.rows_exported == 3
            assert result.output_path.exists()
            assert result.output_path.name == "TEST.json"

            # Read back and verify
            with open(result.output_path) as f:
                data = json.load(f)

            assert data["symbol"] == "TEST"
            assert len(data["data"]) == 3
            assert data["data"][0]["close"] == 104.0

    def test_export_json_with_metadata(self, sample_data: pl.DataFrame) -> None:
        """Test JSON export with metadata."""
        with tempfile.TemporaryDirectory() as tmpdir:
            config = ExportConfig(
                output_path=Path(tmpdir),
                format="json",
                include_metadata=True,
            )

            exporter = JSONExporter(config)
            result = exporter.export(sample_data, "TEST")

            assert result.success

            # Read back and verify
            with open(result.output_path) as f:
                data = json.load(f)

            assert "metadata" in data
            assert data["metadata"]["rows"] == 3
            assert "exported_at" in data["metadata"]

    def test_export_batch_json(self, sample_data: pl.DataFrame) -> None:
        """Test batch JSON export."""
        with tempfile.TemporaryDirectory() as tmpdir:
            config = ExportConfig(
                output_path=Path(tmpdir),
                format="json",
            )

            datasets = {
                "AAPL": sample_data,
                "GOOGL": sample_data.clone(),
            }

            exporter = JSONExporter(config)
            results = exporter.export_batch(datasets)

            assert len(results) == 1  # Single file for batch
            assert results[0].success

            # Read back and verify
            with open(results[0].output_path) as f:
                data = json.load(f)

            assert "AAPL" in data
            assert "GOOGL" in data
            assert len(data["AAPL"]) == 3
            assert len(data["GOOGL"]) == 3


class TestExcelExporter:
    """Test Excel export functionality."""

    @pytest.fixture
    def sample_data(self) -> pl.DataFrame:
        """Create sample OHLCV data."""
        return pl.DataFrame(
            {
                "timestamp": ["2024-01-01", "2024-01-02", "2024-01-03"],
                "open": [100.0, 101.0, 102.0],
                "high": [105.0, 106.0, 107.0],
                "low": [99.0, 100.0, 101.0],
                "close": [104.0, 105.0, 106.0],
                "volume": [1000000, 1100000, 1200000],
            }
        )

    def test_export_excel_basic(self, sample_data: pl.DataFrame) -> None:
        """Test basic Excel export."""
        with tempfile.TemporaryDirectory() as tmpdir:
            config = ExportConfig(
                output_path=Path(tmpdir),
                format="excel",
            )

            exporter = ExcelExporter(config)
            result = exporter.export(sample_data, "TEST")

            assert result.success
            assert result.rows_exported == 3
            assert result.output_path.exists()
            assert result.output_path.suffix == ".xlsx"

    def test_export_excel_with_metadata(self, sample_data: pl.DataFrame) -> None:
        """Test Excel export with metadata."""
        with tempfile.TemporaryDirectory() as tmpdir:
            config = ExportConfig(
                output_path=Path(tmpdir),
                format="excel",
                include_metadata=True,
            )

            exporter = ExcelExporter(config)
            result = exporter.export(sample_data, "TEST")

            assert result.success
            assert result.output_path.exists()

    def test_export_excel_batch(self, sample_data: pl.DataFrame) -> None:
        """Test Excel batch export."""
        with tempfile.TemporaryDirectory() as tmpdir:
            config = ExportConfig(
                output_path=Path(tmpdir) / "batch_export.xlsx",
                format="excel",
            )

            datasets = {
                "AAPL": sample_data,
                "GOOGL": sample_data.clone(),
            }

            exporter = ExcelExporter(config)
            results = exporter.export_batch(datasets)

            assert len(results) == 1
            assert results[0].success
            assert results[0].output_path.exists()

    def test_export_excel_compression(self, sample_data: pl.DataFrame) -> None:
        """Test Excel export with compression."""
        with tempfile.TemporaryDirectory() as tmpdir:
            config = ExportConfig(
                output_path=Path(tmpdir),
                format="excel",
                compression="gzip",
            )

            exporter = ExcelExporter(config)
            result = exporter.export(sample_data, "TEST")

            assert result.success
            assert result.output_path.exists()
            # Compression adds .gz extension
            assert result.output_path.suffix == ".gz"

    def test_export_excel_custom_filename(self, sample_data: pl.DataFrame) -> None:
        """Test Excel export with custom filename."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_file = Path(tmpdir) / "custom_name.xlsx"
            config = ExportConfig(
                output_path=output_file,
                format="excel",
            )

            exporter = ExcelExporter(config)
            result = exporter.export(sample_data, "TEST")

            assert result.success
            assert result.output_path == output_file
            assert result.output_path.exists()


class _InMemoryStorage:
    """Minimal in-memory storage for export tests."""

    def __init__(self):
        self._store: dict[str, DataObject] = {}

    def write(self, data: DataObject) -> str:
        key = f"{data.metadata.asset_class}/{data.metadata.frequency}/{data.metadata.symbol}"
        self._store[key] = deepcopy(data)
        return key

    def read(self, key: str) -> DataObject:
        if key not in self._store:
            raise KeyError(key)
        return deepcopy(self._store[key])

    def exists(self, key: str) -> bool:
        return key in self._store

    def delete(self, key: str) -> None:
        self._store.pop(key, None)

    def list_keys(self, prefix: str = "") -> list[str]:
        return sorted(k for k in self._store if k.startswith(prefix))


class TestExportManager:
    """Test export manager functionality."""

    @pytest.fixture
    def sample_storage(self):
        """Create storage with sample data."""
        from datetime import datetime

        storage = _InMemoryStorage()

        df = pl.DataFrame(
            {
                "timestamp": [
                    datetime(2024, 1, 1),
                    datetime(2024, 1, 2),
                    datetime(2024, 1, 3),
                ],
                "open": [100.0, 101.0, 102.0],
                "high": [105.0, 106.0, 107.0],
                "low": [99.0, 100.0, 101.0],
                "close": [104.0, 105.0, 106.0],
                "volume": [1000000, 1100000, 1200000],
            }
        )

        metadata = Metadata(
            provider="test",
            symbol="AAPL",
            bar_type="time",
            bar_params={"frequency": "daily"},
            asset_class="equities",
            start="2024-01-01",
            end="2024-01-03",
        )

        data_obj = DataObject(data=df, metadata=metadata)
        storage.write(data_obj)

        metadata_googl = Metadata(
            provider="test",
            symbol="GOOGL",
            bar_type="time",
            bar_params={"frequency": "daily"},
            asset_class="equities",
            start="2024-01-01",
            end="2024-01-03",
        )
        data_obj_googl = DataObject(data=df, metadata=metadata_googl)
        storage.write(data_obj_googl)

        return storage

    def test_export_single_dataset(self, sample_storage) -> None:
        """Test exporting a single dataset."""
        with tempfile.TemporaryDirectory() as tmpdir:
            manager = ExportManager(storage=sample_storage)

            result = manager.export(
                key="equities/daily/AAPL",
                output_path=tmpdir,
                format_type="csv",
            )

            assert result.success
            assert result.rows_exported == 3
            assert (Path(tmpdir) / "AAPL.csv").exists()

    def test_export_batch_datasets(self, sample_storage) -> None:
        """Test exporting multiple datasets."""
        with tempfile.TemporaryDirectory() as tmpdir:
            manager = ExportManager(storage=sample_storage)

            results = manager.export_batch(
                keys=["equities/daily/AAPL", "equities/daily/GOOGL"],
                output_path=tmpdir,
                format_type="json",
            )

            assert len(results) == 1  # Single JSON file
            assert results[0].success

    def test_export_pattern(self, sample_storage) -> None:
        """Test exporting datasets matching a pattern."""
        with tempfile.TemporaryDirectory() as tmpdir:
            manager = ExportManager(storage=sample_storage)

            results = manager.export_pattern(
                pattern="equities/daily/*",
                output_path=tmpdir,
                format_type="csv",
            )

            # CSV creates separate files
            assert len(results) == 2
            assert all(r.success for r in results)

    def test_list_formats(self) -> None:
        """Test listing available formats."""
        formats = ExportManager.list_formats()

        assert "csv" in formats
        assert "json" in formats
        assert "excel" in formats
        assert "xlsx" in formats


class TestExportManagerProductionStorage:
    """Exercise export through the production storage protocol."""

    @pytest.fixture(params=[FlatStorage, HiveStorage])
    def storage(self, request, tmp_path):
        storage_class = request.param
        strategy = "flat" if storage_class is FlatStorage else "hive"
        config = StorageConfig(
            base_path=tmp_path / strategy,
            strategy=strategy,
            partition_granularity="day",
        )
        storage = storage_class(config)
        frame = pl.DataFrame(
            {
                "timestamp": [
                    datetime(2024, 1, 2, 14, 30, tzinfo=UTC),
                    datetime(2024, 1, 3, 14, 30, tzinfo=UTC),
                ],
                "open": [100.0, 101.0],
                "high": [105.0, 106.0],
                "low": [99.0, 100.0],
                "close": [104.0, None],
                "volume": [1_000_000.0, 1_100_000.0],
                "label": ["München", "東京"],
            }
        )
        storage.write(
            frame.lazy(),
            "equities/daily/AAPL",
            metadata={"symbol": "AAPL", "provider": "test"},
        )
        storage.write(
            frame.lazy(),
            "equities/hourly/MSFT",
            metadata={"symbol": "MSFT", "provider": "test"},
        )
        return storage

    @pytest.mark.parametrize("format_type", ["csv", "json", "excel"])
    def test_single_export_round_trips_values_and_column_order(
        self, storage, tmp_path, format_type
    ):
        output_dir = tmp_path / f"output-{format_type}"
        output_dir.mkdir()
        manager = ExportManager(storage)
        columns = ["timestamp", "close", "open", "label"]

        result = manager.export(
            "equities/daily/AAPL",
            output_dir,
            format_type,
            columns=columns,
            include_metadata=False,
        )

        assert result.success, result.error
        if format_type == "csv":
            exported = pl.read_csv(result.output_path)
            records = exported.to_dicts()
        elif format_type == "json":
            payload = json.loads(result.output_path.read_text())
            records = payload["data"]
            assert list(records[0]) == columns
        else:
            from openpyxl import load_workbook

            workbook = load_workbook(result.output_path, read_only=True, data_only=True)
            worksheet = workbook["AAPL"]
            rows = list(worksheet.iter_rows(values_only=True))
            assert list(rows[0]) == columns
            records = [dict(zip(rows[0], row, strict=True)) for row in rows[1:]]

        assert list(records[0]) == columns
        assert records[0]["label"] == "München"
        assert records[1]["label"] == "東京"
        assert records[1]["close"] is None
        timestamp = records[0]["timestamp"]
        parsed = timestamp if isinstance(timestamp, datetime) else datetime.fromisoformat(timestamp)
        assert parsed == datetime(2024, 1, 2, 14, 30, tzinfo=UTC)

    def test_pattern_uses_glob_matching_with_zero_argument_list_keys(self, storage, tmp_path):
        output_dir = tmp_path / "pattern"
        output_dir.mkdir()

        results = ExportManager(storage).export_pattern(
            "equities/*/A*", output_dir, "csv", include_metadata=False
        )

        assert len(results) == 1
        assert results[0].success
        assert results[0].output_path.name == "AAPL.csv"

    @pytest.mark.parametrize("format_type", ["csv", "json", "excel"])
    def test_batch_uses_symbols_from_storage_metadata(self, storage, tmp_path, format_type):
        suffix = "xlsx" if format_type == "excel" else format_type
        output_path = tmp_path / ("batch" if format_type == "csv" else f"batch.{suffix}")
        if format_type == "csv":
            output_path.mkdir()

        results = ExportManager(storage).export_batch(
            ["equities/daily/AAPL", "equities/hourly/MSFT"],
            output_path,
            format_type,
            include_metadata=False,
        )

        assert all(result.success for result in results), [result.error for result in results]
        if format_type == "csv":
            assert [result.output_path.name for result in results] == ["AAPL.csv", "MSFT.csv"]
        elif format_type == "json":
            payload = json.loads(output_path.read_text())
            assert list(payload) == ["AAPL", "MSFT"]
        else:
            from openpyxl import load_workbook

            workbook = load_workbook(output_path, read_only=True, data_only=True)
            assert workbook.sheetnames == ["AAPL", "MSFT"]

    def test_missing_requested_column_returns_explicit_failure(self, storage, tmp_path):
        result = ExportManager(storage).export(
            "equities/daily/AAPL",
            tmp_path / "missing.csv",
            "csv",
            columns=["timestamp", "not_present"],
        )

        assert result.success is False
        assert result.error == "Requested export columns are missing: ['not_present']"

    def test_storage_filters_are_pushed_down(self, storage, tmp_path, monkeypatch):
        calls = []
        storage_read = storage.read

        def recording_read(key, start_date=None, end_date=None, columns=None):
            calls.append((start_date, end_date, columns))
            return storage_read(key, start_date, end_date, columns)

        monkeypatch.setattr(storage, "read", recording_read)
        result = ExportManager(storage).export(
            "equities/daily/AAPL",
            tmp_path / "filtered.csv",
            "csv",
            date_filter=("2024-01-02", "2024-01-02"),
            columns=["timestamp", "close"],
            include_metadata=False,
        )

        assert result.success, result.error
        assert calls[-1] == (
            datetime(2024, 1, 2, tzinfo=UTC),
            datetime(2024, 1, 3, tzinfo=UTC),
            ["timestamp", "close"],
        )
        assert pl.read_csv(result.output_path).height == 1

    def test_batch_symbol_collision_is_reported_without_aborting(self, storage, tmp_path):
        frame = storage.read("equities/daily/AAPL").collect()
        storage.write(
            frame,
            "equities/hourly/AAPL-copy",
            metadata={"symbol": "AAPL", "provider": "test"},
        )
        output = tmp_path / "collision"
        output.mkdir()

        results = ExportManager(storage).export_batch(
            ["equities/daily/AAPL", "equities/hourly/AAPL-copy"],
            output,
            "csv",
            include_metadata=False,
        )

        assert len(results) == 2
        assert sum(result.success for result in results) == 1
        assert any("Multiple storage keys" in (result.error or "") for result in results)
        assert (output / "AAPL.csv").is_file()


def test_openpyxl_batch_serializes_non_utc_datetimes(tmp_path, monkeypatch):
    """The runtime openpyxl path accepts aware datetimes and preserves offsets."""
    from openpyxl import load_workbook

    import ml4t.data.export.formats.excel as excel_module

    monkeypatch.setattr(excel_module, "EXCEL_ENGINE", "openpyxl")
    timestamp = pl.Series(
        "timestamp",
        [datetime(2024, 1, 2, 9, 30)],
        dtype=pl.Datetime("us", "America/New_York"),
    )
    exporter = ExcelExporter(
        ExportConfig(
            output_path=tmp_path / "openpyxl.xlsx",
            format="excel",
            include_metadata=False,
        )
    )

    result = exporter.export_batch({"AAPL": pl.DataFrame({"timestamp": timestamp})})

    assert result[0].success, result[0].error
    workbook = load_workbook(result[0].output_path, read_only=True, data_only=True)
    assert workbook["AAPL"]["A2"].value.endswith("-05:00")


def test_hive_export_is_sorted_independently_of_partition_creation_order(tmp_path):
    storage = HiveStorage(
        StorageConfig(
            base_path=tmp_path / "hive-order",
            strategy="hive",
            partition_granularity="day",
        )
    )
    storage.write(
        pl.DataFrame(
            {
                "timestamp": [
                    datetime(2024, 1, 3, tzinfo=UTC),
                    datetime(2024, 1, 2, tzinfo=UTC),
                ],
                "close": [103.0, 102.0],
            }
        ),
        "equities/daily/AAPL",
    )

    result = ExportManager(storage).export(
        "equities/daily/AAPL", tmp_path / "ordered.csv", "csv", include_metadata=False
    )

    assert result.success, result.error
    exported = pl.read_csv(result.output_path, try_parse_dates=True)
    assert exported["close"].to_list() == [102.0, 103.0]
